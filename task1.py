import pandas as pd
import plotly.express as px
from plotly.subplots import make_subplots
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from datetime import datetime
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# Reading data
file_path = '/Users/irminacharchuta/CDQ/Task.xlsx'
df = pd.read_excel(file_path, sheet_name="Outliers Marked")

# Columns to consider for outliers
columns_of_interest = ['mileage', 'price', 'hp', 'year']

# Create the 'output' directory if it doesn't exist
output_dir = 'output'
os.makedirs(output_dir, exist_ok=True)

for i, column in enumerate(columns_of_interest):
    fig_outliers = px.box(df, y=column, title=f'Box Plot for {column}', points="all")
    outlier_image_path = os.path.join(output_dir, f"{column}_outliers.png")
    fig_outliers.write_image(outlier_image_path)

# Finding outliers
def get_outliers(series):
    Q1 = series.quantile(0.25)
    Q3 = series.quantile(0.75)
    IQR = Q3 - Q1
    lower_bound = Q1 - 1.5 * IQR
    upper_bound = Q3 + 1.5 * IQR
    outliers = series[(series < lower_bound) | (series > upper_bound)]
    return outliers

# Replacing outliers with mean values and rounding them
def impute_outliers_mean(series):
    Q1 = series.quantile(0.25)
    Q3 = series.quantile(0.75)
    IQR = Q3 - Q1
    lower_bound = Q1 - 1.5 * IQR
    upper_bound = Q3 + 1.5 * IQR
    mean_value = series.mean()
    replaced_indices = series[(series < lower_bound) | (series > upper_bound)].index
    series_replaced = np.where((series < lower_bound) | (series > upper_bound), mean_value, series)
    series_replaced = np.round(series_replaced)  # Round the replaced values
    return pd.Series(series_replaced, index=series.index), replaced_indices

# Opening the Excel file
wb = openpyxl.load_workbook(file_path)
ws = wb["Outliers Marked"]

# Applying conditional formatting to mark outliers in the 'Outliers Marked' sheet
for column in columns_of_interest:
    outliers = get_outliers(df[column])
    column_letter = get_column_letter(df.columns.get_loc(column) + 1)
    for index, value in outliers.items():
        if not pd.isnull(value):
            cell = ws[f"{column_letter}{index + 2}"]
            cell.fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")

# Saving the changes
wb.save(file_path)

print("Outliers marked in 'Outliers Marked' sheet.")

# Plotting outliers
outliers_sheet = wb.create_sheet("Outliers Plots")

# Generating box plots for each column to visualize outliers
for i, column in enumerate(columns_of_interest):
    fig_outliers = px.box(df, y=column, title=f'Box Plot for {column}', points="all")
    outlier_image_path = os.path.join('output', f"{column}_outliers.png")
    fig_outliers.write_image(outlier_image_path)
    img_outliers = Image(outlier_image_path)
    # Positioning images in two columns
    col_letter = 'A' if i % 2 == 0 else 'J'
    row_number = 1 + (i // 2) * 20
    outliers_sheet.add_image(img_outliers, f'{col_letter}{row_number}')

# Saving the changes
wb.save(file_path)
print(f"Outlier box plots inserted into Excel file {file_path} in sheet 'Outliers Plots'")

# Creating a new sheet for replaced outliers
outliers_replaced_sheet = wb.create_sheet("Outliers Replaced")

# Applying the imputation and rounding the values
replaced_indices_dict = {}
for column in columns_of_interest:
    df[column], replaced_indices = impute_outliers_mean(df[column])
    replaced_indices_dict[column] = replaced_indices

# Writing the modified dataframe to the new sheet
for row in dataframe_to_rows(df, index=False, header=True):
    outliers_replaced_sheet.append(row)

# Applying conditional formatting to mark replaced cells in the new sheet
for column in columns_of_interest:
    replaced_indices = replaced_indices_dict[column]
    column_letter = get_column_letter(df.columns.get_loc(column) + 1)
    for index in replaced_indices:
        cell = outliers_replaced_sheet[f"{column_letter}{index + 2}"]
        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

# Saving the changes
wb.save(file_path)

print("Outliers replaced and marked in 'Outliers Replaced' sheet.")

# Calculating the average age of cars for each year
current_year = datetime.now().year
df['age'] = current_year - df['year']
avg_age_per_year = df.groupby('year')['age'].mean().reset_index()

# Plotting the trend of average age of cars over time
fig = px.line(avg_age_per_year, x='year', y='age', title='Average Age of Cars Over Time', labels={'year': 'Year', 'age': 'Average Age'})
fig.update_layout(title_text="Average Age of Cars Over Time", xaxis_title="Year", yaxis_title="Average Age")

# Saving the figure as an image
image_path = "/Users/irminacharchuta/CDQ/avg_age_trend.png"
fig.write_image(image_path)
print(f"Figure saved as {image_path}")

# Loading the workbook again to insert the image
ws = wb.create_sheet("Cars Trend")  # Insert at position 2
img = Image(image_path)
ws.add_image(img, 'A1')
wb.save(file_path)
print(f"Image inserted into Excel file {file_path} in sheet 'Cars Trend'")

# New section to analyze fuel type trends over the years
df_replaced = pd.read_excel(file_path, sheet_name="Outliers Replaced")

# Grouping by year and fuel type to get the count of each fuel type per year
fuel_trends = df_replaced.groupby(['year', 'fuel']).size().reset_index(name='count')

# Plotting the trend of fuel type over time
fig_fuel = px.line(fuel_trends, x='year', y='count', color='fuel', title='Fuel Type Trends Over Time', labels={'year': 'Year', 'count': 'Count', 'fuel': 'Fuel Type'})
fig_fuel.update_layout(title_text="Fuel Type Trends Over Time", xaxis_title="Year", yaxis_title="Count")

# Saving the fuel type trend figure as an image
fuel_image_path = "/Users/irminacharchuta/CDQ/fuel_trends.png"
fig_fuel.write_image(fuel_image_path)
print(f"Figure saved as {fuel_image_path}")

# Loading the workbook again to insert the fuel type trend image
ws_fuel = wb.create_sheet("Fuel Trends")  # Insert at position 3
img_fuel = Image(fuel_image_path)
ws_fuel.add_image(img_fuel, 'A1')
wb.save(file_path)
print(f"Image inserted into Excel file {file_path} in sheet 'Fuel Trends'")

# Grouping by year and make to get the count of each car make per year
make_trends = df_replaced.groupby(['year', 'make']).size().reset_index(name='count')

# Calculating the total count of each car make across all years
total_make_count = make_trends.groupby('make')['count'].sum().reset_index().sort_values(by='count', ascending=False)

# Plotting the most popular car brands over the years
fig_popular = px.bar(total_make_count, x='make', y='count', title='Most Popular Car Brands Over the Years', labels={'make': 'Car Make', 'count': 'Total Count'})
fig_popular.update_layout(title_text="Most Popular Car Brands Over the Years", xaxis_title="Car Make", yaxis_title="Total Count")

# Saving the most popular car brands figure as an image
popular_image_path = "/Users/irminacharchuta/CDQ/most_popular_car_brands.png"
fig_popular.write_image(popular_image_path)
print(f"Figure saved as {popular_image_path}")

# Loading the workbook again to insert the most popular car brands image
ws_popular = wb.create_sheet("Most Popular Car Brands")  # Insert at position 4
img_popular = Image(popular_image_path)
ws_popular.add_image(img_popular, 'A1')
wb.save(file_path)
print(f"Image inserted into Excel file {file_path} in sheet 'Most Popular Car Brands'")